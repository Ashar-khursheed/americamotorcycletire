import os
import re
import openpyxl
import json
import urllib.request
from PIL import Image
from concurrent.futures import ThreadPoolExecutor

STORAGE_DIR = r"f:\americamotorcycletire\backend\storage\app\public\products"
os.makedirs(STORAGE_DIR, exist_ok=True)

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'

def get_webp_filename(url):
    if not url or not str(url).strip() or str(url).strip() in ('N/A', 'None', 'null'):
        return ''
    url = str(url).strip()
    # Extract base filename from URL
    base_name = url.split('/')[-1].split('?')[0]
    if '.' in base_name:
        name_part = '.'.join(base_name.split('.')[:-1])
    else:
        name_part = base_name
    
    clean_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', name_part)
    if not clean_name:
        clean_name = 'prod_img'
    return f"{clean_name}.webp"

def download_and_convert_to_webp(url):
    if not url or not str(url).strip() or str(url).strip() in ('N/A', 'None', 'null'):
        return None
    url = str(url).strip()
    
    webp_filename = get_webp_filename(url)
    dest_path = os.path.join(STORAGE_DIR, webp_filename)

    # If webp already exists and is non-empty, skip
    if os.path.exists(dest_path) and os.path.getsize(dest_path) > 1000:
        return webp_filename

    # Download source image
    try:
        if url.startswith('http://') or url.startswith('https://'):
            req = urllib.request.Request(url, headers={'User-Agent': USER_AGENT})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = resp.read()
            
            temp_path = dest_path + '.tmp'
            with open(temp_path, 'wb') as f:
                f.write(data)
            
            # Convert to WebP using Pillow
            with Image.open(temp_path) as img:
                img.convert('RGB').save(dest_path, 'WEBP', quality=85)
            
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            print(f"Downloaded & Converted -> {webp_filename}")
            return webp_filename
    except Exception as e:
        print(f"Failed to process {url}: {e}")
        return None

def main():
    print("Collecting all image URLs from fin.xlsx and products.sql...")
    urls = set()

    # 1. fin.xlsx
    if os.path.exists('fin.xlsx'):
        wb = openpyxl.load_workbook('fin.xlsx', data_only=True)
        sheet = wb[wb.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(h).strip() if h else '' for h in rows[0]]
        for row in rows[1:]:
            row_dict = dict(zip(headers, row))
            p_img = str(row_dict.get('Primary Image URL') or '').strip()
            g_imgs = str(row_dict.get('All Image URLs') or '').strip()
            if p_img and p_img.startswith('http'):
                urls.add(p_img)
            if g_imgs:
                for parts in re.split(r'[;,]', g_imgs):
                    p = parts.strip()
                    if p.startswith('http'):
                        urls.add(p)

    # 2. products.sql
    if os.path.exists('products.sql'):
        with open('products.sql', 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        idx = content.find("VALUES")
        if idx != -1:
            sub = content[idx+6:]
            lines = sub.split('\n')
            for line in lines:
                line_s = line.strip()
                if line_s.startswith('('):
                    # extract URLs via regex
                    found = re.findall(r'https?://[^\s\'",\)]+', line_s)
                    for u in found:
                        urls.add(u)

    print(f"Found {len(urls)} unique external image URLs to download & convert to WebP!")

    # Multi-threaded downloading & converting
    with ThreadPoolExecutor(max_workers=10) as executor:
        executor.map(download_and_convert_to_webp, list(urls))

    print("All image downloads & WebP conversions complete!")

if __name__ == '__main__':
    main()
