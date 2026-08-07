import openpyxl
import mysql.connector
import re
import json
import os

def parse_price(val):
    if not val or str(val).strip() in ('N/A', '', 'None'):
        return 0.0
    match = re.search(r'\$?([\d,]+\.?\d*)', str(val))
    if match:
        return float(match.group(1).replace(',', ''))
    return 0.0

def parse_int(val):
    if not val or str(val).strip() in ('N/A', '', 'None'):
        return 0
    try:
        return int(float(str(val).strip()))
    except:
        return 0

def map_vehicle_type(val):
    if not val or not str(val).strip():
        return 'Street Bike'
    v = str(val).strip()
    mapping_explicit = {
        'Street / Sportbike': 'Street Bike',
        'Cruiser / V-Twin': 'Street Bike',
        'Cafe': 'Street Bike',
        'Scooter': 'Street Bike',
        'Adventure / Dual Sport': 'Dirt Bike',
        'Dirt Bike / Off-Road': 'Dirt Bike',
        'UTVATV': 'UTV/ATV',
        'Utvatv': 'UTV/ATV',
        'Adventure / Dual Sport / Cruiser / V-Twin': 'Dirt Bike',
        'Adventure / Dual Sport / Street / Sportbike': 'Dirt Bike',
        'Street / Sportbike / Adventure / Dual Sport': 'Dirt Bike',
        'Cruiser / V-Twin / Adventure / Dual Sport': 'Dirt Bike',
        'Adventure / Dual Sport / Dirt Bike / Off-Road': 'Dirt Bike',
        'Dirt Bike / Off-Road / Adventure / Dual Sport': 'Dirt Bike',
        'Scooter / Street / Sportbike': 'Street Bike',
        'Scooter / Adventure / Dual Sport': 'Dirt Bike',
        'Cruiser / V-Twin / Street / Sportbike': 'Street Bike',
        'Cafe / Cruiser / V-Twin': 'Street Bike',
        'Cafe / Cruiser / V-Twin / Adventure / Dual Sport': 'Dirt Bike',
        'Street / Sportbike / Cruiser / V-Twin': 'Street Bike',
        'Cruiser / V-Twin / Street / Sportbike / Adventure / Dual Sport': 'Dirt Bike',
        'Street / Sportbike / Cruiser / V-Twin / Dirt Bike / Off-Road': 'Dirt Bike',
        'Scooter / Cruiser / V-Twin': 'Street Bike',
        'Cruiser / V-Twin / Utvatv': 'UTV/ATV',
        'Adventure / Dual Sport / Utvatv': 'UTV/ATV'
    }
    if v in mapping_explicit:
        return mapping_explicit[v]
    vl = v.lower()
    if 'utv' in vl or 'atv' in vl:
        return 'UTV/ATV'
    if 'adventure' in vl or 'dual sport' in vl or 'dirt' in vl or 'off-road' in vl:
        return 'Dirt Bike'
    return 'Street Bike'

def slugify(text):
    text = str(text).lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text.strip('-')

def load_env():
    env_vars = {}
    env_path = os.path.join(os.path.dirname(__file__), 'backend', '.env')
    if not os.path.exists(env_path):
        env_path = os.path.join(os.path.dirname(__file__), '.env')
    
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    env_vars[k.strip()] = v.strip().strip('"').strip("'")
    return env_vars

def main():
    env = load_env()
    db_host = os.environ.get('DB_HOST', env.get('DB_HOST', '127.0.0.1'))
    db_port = int(os.environ.get('DB_PORT', env.get('DB_PORT', 3306)))
    db_user = os.environ.get('DB_USERNAME', env.get('DB_USERNAME', 'root'))
    db_pass = os.environ.get('DB_PASSWORD', env.get('DB_PASSWORD', ''))
    db_name = os.environ.get('DB_DATABASE', env.get('DB_DATABASE', 'americamotor'))

    print(f"Connecting to MySQL database '{db_name}' on {db_host}:{db_port}...", flush=True)

    db = mysql.connector.connect(
        host=db_host,
        port=db_port,
        user=db_user,
        password=db_pass,
        database=db_name
    )
    cursor = db.cursor(dictionary=True)

    excel_file = os.path.join(os.path.dirname(__file__), 'fin.xlsx')
    if not os.path.exists(excel_file):
        excel_file = 'fin.xlsx'

    print(f"Loading Excel dataset from '{excel_file}'...", flush=True)
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    # 1. Categories map
    cursor.execute("SELECT id, name FROM categories")
    cat_map = {row['name'].lower(): row['id'] for row in cursor.fetchall()}
    
    # 2. Brands map
    cursor.execute("SELECT id, name FROM brands")
    brand_map = {row['name'].lower(): row['name'] for row in cursor.fetchall()}

    # Clear old tables
    print("Truncating existing product & fitment tables...", flush=True)
    cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
    cursor.execute("TRUNCATE TABLE product_fitments")
    cursor.execute("TRUNCATE TABLE product_attribute_values")
    cursor.execute("TRUNCATE TABLE product_variants")
    cursor.execute("TRUNCATE TABLE product_reviews")
    cursor.execute("TRUNCATE TABLE products")
    cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
    db.commit()

    # Process Sheet 1: All Scraped Products
    sheet1 = wb['All Scraped Products']
    rows1 = list(sheet1.iter_rows(values_only=True))
    headers1 = [str(h).strip() if h else '' for h in rows1[0]]

    product_id_map = {}
    used_slugs = set()

    print(f"Importing {len(rows1)-1} products from Sheet 1...", flush=True)

    for row_idx, row in enumerate(rows1[1:], start=2):
        row_dict = dict(zip(headers1, row))
        p_name = str(row_dict.get('Product Name') or '').strip()
        if not p_name:
            continue

        brand_name = str(row_dict.get('Brand') or 'BMG').strip()
        if brand_name.lower() not in brand_map:
            b_slug = slugify(brand_name)
            cursor.execute("INSERT INTO brands (name, slug, is_active, created_at, updated_at) VALUES (%s, %s, 1, NOW(), NOW())", (brand_name, b_slug))
            brand_map[brand_name.lower()] = brand_name

        cat_name = str(row_dict.get('Category') or 'Tires').strip()
        if cat_name.lower() not in cat_map:
            c_slug = slugify(cat_name)
            cursor.execute("INSERT INTO categories (name, slug, is_active, created_at, updated_at) VALUES (%s, %s, 1, NOW(), NOW())", (cat_name, c_slug))
            cat_id = cursor.lastrowid
            cat_map[cat_name.lower()] = cat_id
        else:
            cat_id = cat_map[cat_name.lower()]

        base_slug = slugify(p_name)
        slug = base_slug
        counter = 1
        while slug in used_slugs:
            slug = f"{base_slug}-{counter}"
            counter += 1
        used_slugs.add(slug)

        raw_vt = str(row_dict.get('Vehicle Type') or '').strip()
        vehicle_type = map_vehicle_type(raw_vt) if raw_vt else 'Street Bike'
        product_type = str(row_dict.get('Specific Product Type') or '').strip() or None
        compatible_makes = str(row_dict.get('Compatible Bike Makes') or '').strip() or None
        compatible_models = str(row_dict.get('Compatible Bike Models') or '').strip() or None
        fitment_year_range = str(row_dict.get('Fitment Year / Range') or '').strip() or None
        item_number = str(row_dict.get('Item Number') or '').strip() or None
        sku = str(row_dict.get('SKU Number') or '').strip() or item_number or f"SKU-{slug}"
        
        price = parse_price(row_dict.get('Retail Price'))
        was_price = parse_price(row_dict.get('Was Price / MSRP'))
        savings = str(row_dict.get('Savings') or '').strip() or None
        rating = float(parse_price(row_dict.get('Rating')) or 0.0)
        review_count = parse_int(row_dict.get('Review Count'))

        front_tire_fitment = str(row_dict.get('Front Tire Fitment') or '').strip() or None
        rear_tire_fitment = str(row_dict.get('Rear Tire Fitment') or '').strip() or None
        wheel_locations = str(row_dict.get('Wheel Locations') or '').strip() or None
        available_sizes_count = parse_int(row_dict.get('Available Sizes Count'))
        available_sizes = str(row_dict.get('Available Sizes') or '').strip() or None
        total_part_numbers = parse_int(row_dict.get('Total Part Numbers'))

        primary_image = str(row_dict.get('Primary Image URL') or '').strip() or None
        all_imgs_raw = str(row_dict.get('All Image URLs') or '').strip()
        gallery_images = []
        if all_imgs_raw:
            gallery_images = [img.strip() for img in all_imgs_raw.split(';') if img.strip()]
        if primary_image and primary_image not in gallery_images:
            gallery_images.insert(0, primary_image)

        desc = str(row_dict.get('Description') or '').strip() or None
        specs = str(row_dict.get('Specs & Features') or '').strip() or None
        fitment_vehicle = str(row_dict.get('Fitment Vehicle') or '').strip() or None
        fitment_disclaimer = str(row_dict.get('Fitment Disclaimer') or '').strip() or None
        source_url = str(row_dict.get('URL') or '').strip() or None

        sql = """
            INSERT INTO products (
                sku, name, slug, brand, category_id, vehicle_type, product_type,
                compatible_makes, compatible_models, fitment_year_range, item_number,
                price, was_price, compare_at_price, cost_price, savings, rating, review_count,
                front_tire_fitment, rear_tire_fitment, wheel_locations, available_sizes_count,
                available_sizes, total_part_numbers, description, specs_and_features,
                fitment_vehicle, fitment_disclaimer, primary_image, gallery_images,
                source_url, stock_quantity, is_active, is_featured, created_at, updated_at
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, 1, %s, NOW(), NOW()
            )
        """
        is_featured = 1 if (row_idx % 5 == 0 or rating >= 4.8) else 0

        cursor.execute(sql, (
            sku, p_name, slug, brand_name, cat_id, vehicle_type, product_type,
            compatible_makes, compatible_models, fitment_year_range, item_number,
            price, was_price, was_price, None, savings, rating, review_count,
            front_tire_fitment, rear_tire_fitment, wheel_locations, available_sizes_count,
            available_sizes, total_part_numbers, desc, specs,
            fitment_vehicle, fitment_disclaimer, primary_image, json.dumps(gallery_images),
            source_url, 50, is_featured
        ))

        pid = cursor.lastrowid
        product_id_map[p_name.lower()] = pid

    db.commit()
    print(f"Inserted {len(product_id_map)} products into 'products' table.", flush=True)

    # Process Fitments from Sheet 3
    sheet3 = wb['Size Breakdown & Fitments']
    rows3 = list(sheet3.iter_rows(values_only=True))
    headers3 = [str(h).strip() if h else '' for h in rows3[0]]

    print(f"Processing fitment relationships from Sheet 3 ({len(rows3)-1} rows)...", flush=True)

    seen_fitment_keys = set()
    fitment_batch = []

    for row in rows3[1:]:
        row_dict = dict(zip(headers3, row))
        p_name = str(row_dict.get('Product Name') or '').strip()
        pid = product_id_map.get(p_name.lower())
        if not pid:
            continue

        makes_str = str(row_dict.get('Compatible Bike Makes') or '').strip()
        models_str = str(row_dict.get('Compatible Bike Models') or '').strip()
        year_str = str(row_dict.get('Fitment Year / Range') or '').strip()
        tire_size = str(row_dict.get('Tire Size') or '').strip() or None
        sku_num = str(row_dict.get('SKU Number') or '').strip() or None
        item_num = str(row_dict.get('Item Number') or '').strip() or None

        makes = [m.strip() for m in makes_str.split(',') if m.strip()] if makes_str else [None]
        models_list = [m.strip() for m in models_str.split(',') if m.strip()] if models_str else [None]
        years = [year_str] if year_str else [None]

        for mk in makes:
            for md in models_list:
                for yr in years:
                    key = (pid, yr, mk, md, tire_size, sku_num)
                    if key in seen_fitment_keys:
                        continue
                    seen_fitment_keys.add(key)
                    fitment_batch.append((
                        pid, yr, mk, md, None, tire_size, sku_num, item_num, sku_num, None
                    ))

    # Also build fallback fitments from Sheet 1 for items without Sheet 3 entries
    cursor.execute("SELECT id, name, compatible_makes, compatible_models, fitment_year_range, item_number, sku FROM products")
    all_prods = cursor.fetchall()
    prods_with_fitments = {k[0] for k in seen_fitment_keys}

    for prod in all_prods:
        pid = prod['id']
        if pid not in prods_with_fitments:
            makes_str = prod['compatible_makes'] or ''
            models_str = prod['compatible_models'] or ''
            year_str = prod['fitment_year_range'] or ''

            makes = [m.strip() for m in makes_str.split(',') if m.strip()] if makes_str else ['Universal']
            models_list = [m.strip() for m in models_str.split(',') if m.strip()] if models_str else ['All Models']
            years = [year_str] if year_str else [None]

            for mk in makes:
                for md in models_list:
                    for yr in years:
                        key = (pid, yr, mk, md, None, prod['sku'])
                        if key in seen_fitment_keys:
                            continue
                        seen_fitment_keys.add(key)
                        fitment_batch.append((
                            pid, yr, mk, md, None, None, prod['sku'], prod['item_number'], prod['sku'], None
                        ))

    fitment_insert_sql = """
        INSERT INTO product_fitments (
            product_id, year, make, model, position, tire_size, sku_number, item_number, vendor_part_number, notes, created_at, updated_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
    """

    print(f"Batch inserting {len(fitment_batch)} fitments...", flush=True)

    cursor_fast = db.cursor()
    chunk_size = 2000
    for i in range(0, len(fitment_batch), chunk_size):
        chunk = fitment_batch[i:i+chunk_size]
        cursor_fast.executemany(fitment_insert_sql, chunk)
        db.commit()

    print(f"\n==========================================", flush=True)
    print(f" SUCCESS! Live Database Import Completed!", flush=True)
    print(f" Total Products: {len(product_id_map)}", flush=True)
    print(f" Total Fitments: {len(fitment_batch)}", flush=True)
    print(f"==========================================\n", flush=True)

    cursor_fast.close()
    cursor.close()
    db.close()

if __name__ == '__main__':
    main()
